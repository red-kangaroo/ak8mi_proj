# !/usr/bin/env python3
# -*- coding: utf-8 -*-

import datetime
import logging as log
import math
import matplotlib.pyplot as plt
from openpyxl import Workbook
from pathlib import Path
import random
import statistics
import sys

"""
Stochastic heuristic algorithms

Random Search and Hill Climber for: 
* 1st DeJong function
* 2nd DeJong function
* Schwefel function

@author: Filip Findura
"""

# Requested dimensions, max energy function calls and iterations:
D = [5, 10, 20]
FES = 10000
ITERATIONS = 30

# Hill climber:
H_COUNT = 10  # Number of neighbours
H_NEAR = 0.1  # Max difference of neighbourhood

# Logging root handler:
logger = log.getLogger()
# Logging level:
LOG_LVL = 'INFO'

# Output Excel workbook:
wb = Workbook(write_only=True)

# All results of optimization:
FUNCTIONS = ['de_jong_1', 'de_jong_2', 'schwefel']
RESULTS = dict()


##################################################
# Energy Functions
##################################################
def de_jong_1(inputs: list):
    # http://profesores.elo.utfsm.cl/~tarredondo/info/soft-comp/functions/node2.html
    return sum([x**2 for x in inputs])


def de_jong_2(x: list):
    # http://profesores.elo.utfsm.cl/~tarredondo/info/soft-comp/functions/node5.html
    result = 0

    for i in range(len(x) - 1):
        result += 100 * (x[i]**2 - x[i+1]) ** 2 + (1 - x[i]) ** 2

    return result


def schwefel(x: list):
    # https://www.cs.unm.edu/~neal.holts/dga/benchmarkFunction/schwefel.html
    result = 0
    alpha = 418.982887

    for i in range(len(x)):
        result -= x[i] * math.sin(math.sqrt(math.fabs(x[i])))

    return result + alpha * len(x)


##################################################
# Algorithms
##################################################
def random_search(it: int, fn, dim: int, constr_min: float, constr_max: float):
    logger.info(f"Starting random search for D{dim} {fn.__name__}...")
    ws = wb.create_sheet()
    ws.title = f"Random search {fn.__name__.replace('_', ' ').title()} D{dim}"
    ws.append(["Iteration", "Minimum", "Inputs"])
    plt.figure(1+dim)

    RESULTS[dim]['random_search'][fn.__name__] = [list() for i in range(it)]
    for i in range(it):
        result_in = None
        result_out = None
        graph_x = list()
        graph_y = list()

        for a in range(FES):
            ins = [random.uniform(constr_min, constr_max) for b in range(dim)]
            out = fn(ins)

            if result_out is None or out < result_out:
                result_in = ins
                result_out = out

            graph_x.append(a)
            graph_y.append(result_out)
            RESULTS[dim]['random_search'][fn.__name__][i].append((a, result_out))

        plt.plot(graph_x, graph_y, label=f"Iter. No. {i+1}")
        ws.append([i+1, result_out] + result_in)
        logger.debug(f"For {dim} dimensions: minimum {result_out} with inputs {result_in}.")

    plt.xlabel('Cost Function Evaluations')
    plt.ylabel('Results')
    plt.title(f"All iterations of random search for {fn.__name__.replace('_', ' ').title()}, D{dim}")

    splt = plt.subplot()
    box = splt.get_position()
    splt.set_position([box.x0, box.y0, box.width * 0.8, box.height])
    splt.legend(loc='upper left', bbox_to_anchor=(1, 1))

    plt.show()
    plt.clf()


def hill_climber(it: int, fn, dim: int, constr_min: float, constr_max: float):
    logger.info(f"Starting hill climber for D{dim} {fn.__name__}...")
    ws = wb.create_sheet()
    ws.title = f"Hill climber {fn.__name__.replace('_', ' ').title()} D{dim}"
    ws.append(["Iteration", "Minimum", "Inputs"])
    plt.figure(2+dim)

    def climb_step(ins: list) -> (float, list):
        # Finds the best neighbour.
        nonlocal constr_min, constr_max, fn

        nears = list()
        results = list()
        result = None

        for c in range(H_COUNT):
            nears.append(list())
            for inp in ins:
                # Sadly, Python does not have a do-while loop. I miss it in places like this.
                new_inp = random.uniform(inp * (1 - H_NEAR), inp * (1 + H_NEAR))
                while new_inp > constr_max or new_inp < constr_min:
                    new_inp = random.uniform(inp * (1 - H_NEAR), inp * (1 + H_NEAR))

                nears[c].append(new_inp)

            r = fn(nears[c])
            results.append(r)
            if result is None or r <= results[result]:
                result = c

        return results[result], nears[result]

    def climb_main(uses: int, itr: int):
        nonlocal constr_min, constr_max, dim, fn, graph_x, graph_y, res_out

        first_in = [random.uniform(constr_min, constr_max) for b in range(dim)]
        result_out, result_in = climb_step(first_in)
        uses -= 10

        if res_out is None or result_out < res_out:
            graph_x.append(FES - uses)
            graph_y.append(result_out)
            RESULTS[dim]['hill_climber'][fn.__name__][i].append((FES - uses, result_out))
        else:
            RESULTS[dim]['hill_climber'][fn.__name__][i].append((FES - uses, res_out))

        while uses > 0:
            new_out, new_in = climb_step(result_in)
            uses -= 10

            if new_out >= result_out:
                logger.debug(f"Ending hill climbing for iteration {itr} with {uses} uses still to go.")

                if res_out is None or result_out < res_out:
                    RESULTS[dim]['hill_climber'][fn.__name__][i].append((FES - uses, result_out))
                else:
                    RESULTS[dim]['hill_climber'][fn.__name__][i].append((FES - uses, res_out))

                break
            else:
                result_out, result_in = new_out, new_in

                if res_out is None or result_out < res_out:
                    graph_x.append(FES - uses)
                    graph_y.append(result_out)
                    RESULTS[dim]['hill_climber'][fn.__name__][i].append((FES - uses, result_out))
                else:
                    RESULTS[dim]['hill_climber'][fn.__name__][i].append((FES - uses, res_out))

        return result_out, result_in, uses

    RESULTS[dim]['hill_climber'][fn.__name__] = [list() for i in range(it)]
    for i in range(it):
        res_out = None
        graph_x = list()
        graph_y = list()

        res_out, res_in, uses_left = climb_main(FES, i)
        climbed = 1

        while uses_left > 0:
            logger.debug(f"Climbing again as we have {uses_left} uses left for iteration {i}.")
            new_res_out, new_res_in, uses_left = climb_main(uses_left, i)

            if new_res_out < res_out:
                res_out, res_in = new_res_out, new_res_in

            climbed += 1

        graph_x.append(FES)
        graph_y.append(res_out)
        plt.plot(graph_x, graph_y, label=f"Iter. No. {i+1}")

        ws.append([i+1, res_out] + res_in)
        logger.debug(f"For {dim} dimensions: minimum {res_out} with inputs {res_in}. Climbed {climbed} time(s).")

    plt.xlabel('Cost Function Evaluations')
    plt.ylabel('Results')
    plt.title(f"All iterations of hill climber for {fn.__name__.replace('_', ' ').title()}, D{dim}")

    splt = plt.subplot()
    box = splt.get_position()
    splt.set_position([box.x0, box.y0, box.width * 0.8, box.height])
    splt.legend(loc='upper left', bbox_to_anchor=(1, 1))

    plt.show()
    plt.clf()


def simulated_annealing(it: int, fn, dim: int, constr_min: float, constr_max: float):
    logger.info(f"Starting simulated annealing for D{dim} {fn.__name__}...")
    ws = wb.create_sheet()
    ws.title = f"Simulated annealing {fn.__name__.replace('_', ' ').title()} D{dim}"
    ws.append(["Iteration", "Minimum", "Inputs"])
    plt.figure(3+dim)

    # TODO

    plt.xlabel('Cost Function Evaluations')
    plt.ylabel('Results')
    plt.title(f"All iterations of simulated annealing for {fn.__name__.replace('_', ' ').title()}, D{dim}")

    splt = plt.subplot()
    box = splt.get_position()
    splt.set_position([box.x0, box.y0, box.width * 0.8, box.height])
    splt.legend(loc='upper left', bbox_to_anchor=(1, 1))

    plt.show()
    plt.clf()


##################################################
# Helpers
##################################################
def set_logging(l_name: str, level: str):
    """Set up attributes of the root logger

    :param l_name: string name of the logger
    :param level: string name of base log level
    """
    global logger

    log_format = "%(asctime)s | %(levelname)-5s | %(message)s"
    if level == "DEBUG":
        log_format += " | %(filename)s@ln%(lineno)d"
    formatter = log.Formatter(log_format)

    logger = log.getLogger(l_name)
    logger.setLevel(level)
    handler = log.StreamHandler(sys.stdout)
    handler.setFormatter(formatter)
    logger.addHandler(handler)


def get_results(it: int):
    for d in D:
        RESULTS[d] = {
            'random_search': dict(),
            'hill_climber': dict(),
            # 'simulated_annealing': dict(),
        }
        random_search(it, de_jong_1, d, -5.0, 5.0)
        random_search(it, de_jong_2, d, -5.0, 5.0)
        random_search(it, schwefel, d, -500, 500)

        hill_climber(it, de_jong_1, d, -5.0, 5.0)
        hill_climber(it, de_jong_2, d, -5.0, 5.0)
        hill_climber(it, schwefel, d, -500, 500)

        # TODO: simulated_annealing

        for idx, f in enumerate(FUNCTIONS):
            avg_x = dict()
            med_x = dict()
            max_x = dict()
            min_x = dict()
            graph_y = dict()

            for ndx, k in enumerate(RESULTS[d].keys()):
                plt.figure(d + idx*4 + ndx + 100)
                avg_x[k] = list()
                med_x[k] = list()
                max_x[k] = list()
                min_x[k] = list()
                graph_y[k] = list()

                if k == 'hill_climber':
                    steps = FES // H_COUNT
                else:
                    steps = FES

                for x in range(steps):
                    x_total = list()
                    y_total = 0
                    try:
                        for i in range(it):
                            x_total.append(RESULTS[d][k][f][i][x][1])
                            if RESULTS[d][k][f][i][x][0] > y_total:
                                y_total = RESULTS[d][k][f][i][x][0]
                    except IndexError:
                        logger.error(f"Alg: {k}; fn: {f}; step: {x}/{steps}")
                        sys.exit(1)
                    avg_x[k].append(statistics.mean(x_total))
                    med_x[k].append(statistics.median(x_total))
                    max_x[k].append(max(x_total))
                    min_x[k].append(min(x_total))
                    graph_y[k].append(y_total)

                # I mixed up x and y, so graph_y is axis x and foo_x is on axis y...
                plt.plot(graph_y[k], avg_x[k], label="Average")
                plt.plot(graph_y[k], med_x[k], label="Median")
                plt.plot(graph_y[k], max_x[k], label="Max")
                plt.plot(graph_y[k], min_x[k], label="Min")

                plt.xlabel('Cost Function Evaluations')
                plt.ylabel('Results')
                plt.title(f"Statistics for {k.replace('_', ' ')} for {f.replace('_', ' ').title()}, D{d}")

                splt = plt.subplot()
                box = splt.get_position()
                splt.set_position([box.x0, box.y0, box.width * 0.8, box.height])
                splt.legend(loc='upper left', bbox_to_anchor=(1, 1))

                plt.show()

            plt.figure(idx + 333 + d)
            for k in RESULTS[d].keys():
                plt.plot(graph_y[k], avg_x[k], label=k.replace('_', ' ').title())

            plt.xlabel('Cost Function Evaluations')
            plt.ylabel('Results')
            plt.title(f"Algorithm comparison for {f.replace('_', ' ').title()}, D{d}")

            splt = plt.subplot()
            box = splt.get_position()
            splt.set_position([box.x0, box.y0, box.width * 0.8, box.height])
            splt.legend(loc='upper left', bbox_to_anchor=(1, 1))

            plt.show()

    wb.save(Path("./raw_output.xlsx"))
    wb.close()


if __name__ == "__main__":
    time_start = datetime.datetime.now()

    name = "ak8mi"
    set_logging(name, LOG_LVL)
    get_results(ITERATIONS)

    time_stop = datetime.datetime.now()
    time_elapsed = time_stop - time_start
    logger.info(f"Done {ITERATIONS} iterations over 3 functions and 2 algorithms in {time_elapsed.total_seconds():.3f} "
                f"seconds ({time_elapsed.total_seconds() / 60:.3f} minutes).")
